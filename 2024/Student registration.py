from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
background="#06283D"     
framebg="#EDEDED"
framefg="#06283D"

root=Tk()
root.title("P4L Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)
file_path = 'Student_data.xlsx'
file = pathlib.Path(file_path)

if file.exists():
    pass
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Dob"
    sheet['F1'] = "Date of registration"
    sheet['G1'] = "Religion"
    sheet['H1'] = "Course"
    sheet['I1'] = "Father Name"
    sheet['J1'] = "Mother Name"
    sheet['K1'] = "Father's Occupation"
    sheet['L1'] = "Mother's Occupation"
    workbook.save(filename=file_path)


#################Exit window #############
def Exit():
    root.destroy()

#############Show image  ####################
def showimage():
    global filename
    global img

    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select image file",
                                          filetypes=(("JPG File", "*.jpg"),
                                                     ("PNG File", "*.png"),
                                                     ("All files", "*.*")))

    img = Image.open(filename)
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2

#################Registration No.###########
####Automatic Registration number 

def registration_no():
    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value
   
    try:
        Registration.set(max_row_value+1)

    except:
        Registration.set("1")

##########Clear################
def Clear():
    global img
    Name.set('')
    DOB.set('')
    Course.set('')
    F_Name.set('')
    M_Name.set('')
    Father_occupation.set('')
    Mother_occupation.set('')
    Religion.set('Select Religion')
    Class.set("Select Class")

    registration_no()

    saveButton.config(state='normal')

    img1 = PhotoImage(file='2024/Images/upload.png')
    lbl.config(image=img1)
    lbl.image = img1

    img = ""



############Save
def Save():
  R1=Registration.get()
  N1=Name.get()
  C1=Class.get()
  try:
      G1=gender
  except:
      messagebox.showerror("error","Select Your Gender!")

  D2=DOB.get()
  D1=Date.get()
  Re1=Religion.get()
  Co1=Course.get()
  fathername=F_Name.get()
  mothername=M_Name.get()
  F1=Father_occupation.get()
  M1=Mother_occupation.get()

  if N1=="" or C1=="Select Class" or D2=="" or Re1=="Select Religion" or Co1=="" or fathername=="" or mothername=="" or F1=="" or M1=="":
    messagebox.showerror("error", "Some Data is missing!")

  else:
    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=R1)
    sheet.cell(column=2,row=sheet.max_row,value=N1)
    sheet.cell(column=3,row=sheet.max_row,value=C1)
    sheet.cell(column=4,row=sheet.max_row,value=G1)
    sheet.cell(column=5,row=sheet.max_row,value=D2)
    sheet.cell(column=6,row=sheet.max_row,value=D1)
    sheet.cell(column=7,row=sheet.max_row,value=Re1)
    sheet.cell(column=8,row=sheet.max_row,value=Co1)
    sheet.cell(column=9,row=sheet.max_row,value=fathername)
    sheet.cell(column=10,row=sheet.max_row,value=mothername)
    sheet.cell(column=11,row=sheet.max_row,value=F1)
    sheet.cell(column=12,row=sheet.max_row,value=M1)

    file.save(r'Student_data.xlsx')

    try:
        img.save("Students Images/"+str(R1)+".jpg")
    except:
        messagebox.showinfo("info","Profile Picture is not available!!!")
    
    messagebox.showinfo("info","Successfully data entered!!!")

    Clear()         #Clear entry box and image section

    registration_no()  #It will recheck registration no and reissue new one

   



#################Search#################
def perform_search():
    text = search_box.get()  # Get the registration number from the search box
    Clear()  # Clear the current entries
    saveButton.config(state='disabled')  # Disable save button during search

    try:
        file = openpyxl.load_workbook("Student_data.xlsx")
        sheet = file.active

        name_found = False

        for row in sheet.iter_rows(min_row=2):  # Skip header row
            if row[0].value == int(text):
                Registration.set(row[0].value)  # Set the Registration number
                Name.set(row[1].value)  # Set the Name
                Class.set(row[2].value)  # Set the Class
                gender = row[3].value  # Set the Gender
                if gender == "Male":
                    radio.set(1)
                else:
                    radio.set(2)
                DOB.set(row[4].value)  # Set the Date of Birth
                Date.set(row[5].value)  # Set the Date of registration
                Religion.set(row[6].value)  # Set the Religion
                Course.set(row[7].value)  # Set the Course
                F_Name.set(row[8].value)  # Set the Father's Name
                M_Name.set(row[9].value)  # Set the Mother's Name
                Father_occupation.set(row[10].value)  # Set Father's Occupation
                Mother_occupation.set(row[11].value)  # Set Mother's Occupation
                
                # Load and display the image
                img_path = f"Students Images/{row[0].value}.jpg"
                if os.path.exists(img_path):
                    img = Image.open(img_path)
                    resized_image = img.resize((190, 190))
                    photo2 = ImageTk.PhotoImage(resized_image)
                    lbl.config(image=photo2)
                    lbl.image = photo2
                else:
                    messagebox.showinfo("Info", "Profile Picture is not available!")
                
                name_found = True
                break  # Stop searching once found

        if not name_found:
            raise ValueError("Invalid registration Number")

    except ValueError as e:
        messagebox.showerror("Invalid", str(e))
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")



###############Update####################
def Update():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    selection()
    G1 = gender
    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    Co1 = Course.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_occupation.get()
    M1 = Mother_occupation.get()
   
    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet = file.active

    # Flag to check if registration number is found
    reg_no_found = False

    for row in sheet.iter_rows(min_row=2):  # Start from row 2 to skip headers
        if row[0].value == R1:  # Compare with the value in the first column
            reg_no_position = row[0].row
            reg_no_found = True
            break

    if not reg_no_found:
        messagebox.showerror("Error", "Registration number not found!")
        return

    # Update the data in the Excel sheet
    sheet.cell(column=1, row=reg_no_position, value=R1)
    sheet.cell(column=2, row=reg_no_position, value=N1)
    sheet.cell(column=3, row=reg_no_position, value=C1)
    sheet.cell(column=4, row=reg_no_position, value=G1)
    sheet.cell(column=5, row=reg_no_position, value=D2)
    sheet.cell(column=6, row=reg_no_position, value=D1)
    sheet.cell(column=7, row=reg_no_position, value=Re1)
    sheet.cell(column=8, row=reg_no_position, value=Co1)
    sheet.cell(column=9, row=reg_no_position, value=fathername)
    sheet.cell(column=10, row=reg_no_position, value=mothername)
    sheet.cell(column=11, row=reg_no_position, value=F1)
    sheet.cell(column=12, row=reg_no_position, value=M1)

    # Save the workbook after updating
    file.save(r'Student_data.xlsx')

    try:
        img.save(f"Student Images/{R1}.jpg")
    except:
        pass

    messagebox.showinfo("Update", "Update Successful")
    Clear()  # Clear all entry boxes after update



# Top frames and anchor E to place email at the corner
Label(root, text="Email: 22514@student.embuni.ac.ke", width=10, height=3, bg="#06283D", fg='white', anchor='e').pack(side=TOP, fill='x')
Label(root, text="P4L STUDENT REGISTRATION", width=10, height=2, bg="#FF0000", fg='#ffffff', font='arial 20 bold').pack(side=TOP, fill='x')

# Search box to update
search_box = StringVar()
Entry(root, textvariable=search_box, width=15, bd=2, font="arial 20").place(x=830, y=70)
imageicon3 = PhotoImage(file="2024/Images/search.png")
Srch = Button(root, text="Search", compound=LEFT, image=imageicon3, width=123, bg='#68ddfa', font="arial 13 bold", command=perform_search)
Srch.place(x=1070, y=73)

# Label to display search results
search_result_label = Label(root, text="", font="arial 12", bg=framebg, fg=framefg)
search_result_label.place(x=830, y=130)


imageicon4 = PhotoImage(file="2024/Images/Refresh.png")
update_button = Button(root, image=imageicon4, bg="#c36464", command=Update)
update_button.place(x=20, y=64)

#Registration and date
Label(root,text="Registration No:",font="arial 13",fg=framebg,bg=background).place(x=30,y=150)
Label(root,text="Date:",font="arial 13",fg=framebg,bg=background).place(x=500,y=150)

Registration=IntVar()
Date = StringVar()

reg_entry = Entry(root,textvariable=Registration,width=15,font="arial 10")
reg_entry.place(x=160,y=150)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%y")
date_entry = Entry(root,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=550,y=150)

Date.set(d1)

#Student details
obj=LabelFrame(root,text="Student's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date of Birth:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)

Label(obj,text="Class:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj,text="Religion:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj,text="Course:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)


Name=StringVar()
name_entry = Entry(obj,textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160,y=50)


DOB=StringVar()
dob_entry = Entry(obj,textvariable=DOB,width=20,font="arial 10")
dob_entry.place(x=160,y=100)




# Gender selection function
def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "Female"

# Define the Radiobuttons
radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)

# Place the Radiobuttons
R1.place(x=150, y=150)
R2.place(x=200, y=150)







Religion=Combobox(obj,values=['Christian ','SDA','Muslim','Hindu','Pagan'],font="Roboto 10",width=17,state="r")
Religion.place(x=630,y=100)
Religion.set("Select Religion")



Course=StringVar()
course_entry = Entry(obj,textvariable=Course,width=20,font="arial 10")
course_entry.place(x=630,y=150)

Class= Combobox(obj,values=['Y1','Y2','Y3','Y4','Post Graduate'],font="Roboto 10",width=17,state="r")
Class.place(x=630,y=50)
Class.set("Select Class")



#Parents details
obj2=LabelFrame(root,text="Student's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)

Label(obj2,text="Father's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)


F_Name=StringVar()
f_entry = Entry(obj2,textvariable=F_Name,width=20,font="arial 10")
f_entry.place(x=160,y=50)

Father_occupation=StringVar()
FO_entry = Entry(obj2,textvariable=Father_occupation,width=20,font="arial 10")
FO_entry.place(x=160,y=100)


Label(obj2,text="Mother's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)


M_Name=StringVar()
m_entry = Entry(obj2,textvariable=M_Name,width=20,font="arial 10")
m_entry.place(x=630,y=50)

Mother_occupation=StringVar()
MO_entry = Entry(obj2,textvariable=Mother_occupation,width=20,font="arial 10")
MO_entry.place(x=630,y=100)

#Image
f=Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img=PhotoImage(file="2024/Images/upload.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)


#button
Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=showimage).place(x=1000, y=370)

saveButton=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightgreen", command=Save)
saveButton.place(x=1000,y=450)



Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="lightpink",command=Clear).place(x=1000,y=530)

Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="grey",command=Exit).place(x=1000,y=610)







root.mainloop()